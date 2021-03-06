from tkinter import *
from tkinter.ttk import *

from selenium import webdriver
from selenium.webdriver.chrome.options import Options

from bs4 import BeautifulSoup

import re

import datetime

import openpyxl



dict_regions_russia = {
	'Планета Земля':-1,
	'Без учета региона': 0,
	'Европа': 111,
	'СНГ': 166,
	'Универсальное': 318,
	'Азия': 183,
	'Россия': 225,
	'Северо-Западный федеральный округ': 17,
	'Калининградская область': 10857,
	'Калининград': 22,
	'Мурманская область': 10897,
	'Мурманск': 24,
	'Республика Карелия': 10933,
	'Петрозаводск': 18,
	'Санкт-Петербург и Ленинградская область': 10174,
	'Санкт-Петербург' :2,
	'Псковская область' :10926,
	'Псков' :25,
	'Великие Луки' :10928,
	'Новгородская область' :10904,
	'Великий Новгород' :24,
	'Центральный федеральный округ' :3,
	'Тверская область' :10819,
	'Тверь' :14,
	'Смоленская область' :10795,
	'Смоленск' :12,
	'Брянская область' :10650,
	'Брянск' :191,
	'Калужская область' :10693,
	'Калуга' :6,
	'Обнинск' :967,
	'Курская область' :10705,
	'Курск' :8,
	'Орловская область' :10772,
	'Орёл' :10,
	'Тульская область' :10832,
	'Тула' :15,
	'Москва и Московская область' :1,
	'Москва' :213,
	'Долгопрудный' :214,
	'Дубна' :215,
	'Зеленоград' :216,
	'Пущино' :217,
	'Белгородская область' :10645,
	'Белгород' :4,
	'Липецкая область' :10712,
	'Липецк' :9,
	'Ярославская область' :10841,
	'Ярославль' :16,
	'Владимирская область' :10658,
	'Владимир' :192,
	'Александров' :10656,
	'Гусь-Хрустальный' :10661,
	'Муром' :10668,
	'Ивановская область' :10687,
	'Иваново' :5,
	'Рязанская область' :10776,
	'Рязань' :11,
	'Тамбовская область' :10802,
	'Тамбов' :13,
	'Воронежская область' :10672,
	'Воронеж' :193,
	'Южный федеральный округ' :26,
	'Ростовская область' :11029,
	'Ростов-на-Дону' :39,
	'Шахты' :11053,
	'Таганрог' :971,
	'Новочеркасск' :238,
	'Волгодонск' :11036,
	'Краснодарский край' :10995,
	'Краснодар' :35,
	'Анапа' :1107,
	'Новороссийск' :970,
	'Сочи' :239,
	'Туапсе' :1058,
	'Геленджик' :10990,
	'Армавир' :10987,
	'Ейск' :10993,
	'Республика Адыгея' :11004,
	'Майкоп' :1093,
	'Карачаево-Черкесская республика' :11020,
	'Черкесск' :1104,
	'Кабардино-Балкарская республика' :11013,
	'Нальчик' :30,
	'Северная Осетия' :11021,
	'Владикавказ' :33,
	'Республика Ингушетия' :11012,
	'Чеченская республика' :11024,
	'Грозный' :1106,
	'Республика Дагестан' :11010,
	'Махачкала' :28,
	'Ставропольский край' :11069,
	'Ставрополь' :36,
	'Каменск-Шахтинский' :11043,
	'Пятигорск' :11067,
	'Минеральные Воды' :11063,
	'Ессентуки' :11057,
	'Кисловодск' :11062,
	'Республика Калмыкия' :11015,
	'Элиста' :1094,
	'Астраханская область' :10946,
	'Астрахань' :37,
	'Волгоградская область' :10950,
	'Волгоград' :38,
	'Поволжье' :40,
	'Саратовская область' :11146,
	'Саратов' :194,
	'Жигулевск' :11132,
	'Балаково' :11143,
	'Пензенская область' :11095,
	'Пенза' :49,
	'Республика Мордовия' :11117,
	'Саранск' :42,
	'Ульяновская область' :11153,
	'Ульяновск' :195,
	'Самарская область' :11131,
	'Самара' :51,
	'Тольятти' :240,
	'Сызрань' :11139,
	'Чувашская республика' :11156,
	'Чебоксары' :45,
	'Республика Марий Эл' :11077,
	'Йошкар-Ола' :41,
	'Нижегородская область' :11079,
	'Нижний Новгород' :47,
	'Саров' :11083,
	'Кировская область' :11070,
	'Киров' :46,
	'Костромская область' :10699,
	'Кострома' :7,
	'Вологодская область' :10853,
	'Вологда' :21,
	'Архангельская область' :10842,
	'Архангельск' :20,
	'Северодвинск' :10849,
	'Ненецкий автономный округ' :10176,
	'Республика Коми' :10939,
	'Сыктывкар' :19,
	'Удмуртская республика' :11148,
	'Ижевск' :44,
	'Республика Татарстан' :11119,
	'Казань' :43,
	'Набережные Челны' :236,
	'Нижнекамск' :11127,
	'Пермский край' :11108,
	'Пермь' :50,
	'Республика Башкортостан' :11111,
	'Уфа' :172,
	'Нефтекамск' :11114,
	'Салават' :11115,
	'Стерлитамак' :11116,
	'Оренбургская область' :11084,
	'Оренбург' :48,
	'Дзержинск' :972,
	'Урал' :52,
	'Челябинская область' :11225,
	'Челябинск' :56,
	'Магнитогорск' :235,
	'Снежинск' :11218,
	'Курганская область' :11158,
	'Курган' :53,
	'Свердловская область' :11162,
	'Екатеринбург' :54,
	'Каменск-Уральский' :11164,
	'Нижний Тагил' :11168,
	'Новоуральск' :11170,
	'Первоуральск' :11171,
	'Тюменская область' :11176,
	'Тюмень' :55,
	'Тобольск' :11175,
	'Ханты-Мансийский автономный округ' :11193,
	'Ханты-Мансийск' :57,
	'Сургут' :973,
	'Нижневартовск' :1091,
	'Сибирь' :59,
	'Омская область' :11318,
	'Омск' :66,
	'Новосибирская область' :11316,
	'Новосибирск' :65,
	'Бердск' :11314,
	'Томская область' :11353,
	'Томск' :67,
	'Ямало-Ненецкий автономный округ' :11232,
	'Салехард' :58,
	'Алтайский край' :11235,
	'Барнаул' :197,
	'Бийск' :975,
	'Рубцовск' :11251,
	'Республика Алтай' :10231,
	'Горно-Алтайск' :11319,
	'Кемеровская область' :11282,
	'Кемерово' :64,
	'Междуреченск' :11287,
	'Новокузнецк' :237,
	'Прокопьевск' :11291,
	'Республика Хакасия' :11340,
	'Абакан' :1095,
	'Республика Тыва' :10233,
	'Кызыл' :11333,
	'Красноярский край' :11309,
	'Красноярск' :62,
	'Ачинск' :11302,
	'Норильск' :11311,
	'Железногорск' :20086,
	'Иркутская область' :11266,
	'Иркутск' :63,
	'Братск' :976,
	'Республика Бурятия' :11330,
	'Улан-Удэ' :198,
	'Забайкальский край' :21949,
	'Чита' :68,
	'Дальневосточный федеральный округ' :73,
	'Республика Саха (Якутия)' :11443,
	'Якутск' :74,
	'Амурская область' :11375,
	'Благовещенск' :77,
	'Еврейская автономная область' :10243,
	'Биробиджан' :11393,
	'Приморский край' :11409,
	'Владивосток' :75,
	'Находка' :974,
	'Уссурийск' :11426,
	'Чукотский автономный округ' :10251,
	'Анадырь' :11458,
	'Камчатский край' :11398,
	'Петропавловск-Камчатский' :78,
	'Магаданская область' :11403,
	'Магадан' :79,
	'Сахалинская область' :11450,
	'Южно-Сахалинск' :80,
	'Хабаровский край' :11457,
	'Хабаровск' :76,
	'Комсомольск-на-Амуре' :11453
}

exceptions = [
	'2gis.ru',
	'yandex.ru',
	'wikipedia',
	'pulscen',
	'blizko.ru',
	'Avito.ru',
	'avito.ru',
	'edadeal.ru'
]

def stop_pars(event):
	print('Stop Pars')

def output(event):

	button_1.config(state="disabled")
	# button_2.config(state="normal")
	root.update()

# получаю список запросов
	inquiries_text = text.get(1.0, END)
	inquiries_text = inquiries_text.split("\n")
	inquiries = []
	for val in inquiries_text:
		if len(val) != 0:
			inquiries.append(val.strip())
# получаю регион
	region = dict_regions_russia[combo.get()]
# получаю глубину парсинга
	deep_pars = spin.get()
	try:
		deep_pars = int(deep_pars)
	# отрабатывает исклчюение на то что ввели не цифру
	except ValueError:
		deep_pars = 1
	# максимальная глубина парсинга 10 страниц
	if deep_pars > 10:
		deep_pars = 10
	# если ввели ноль
	if deep_pars == 0:
		deep_pars = 1

	progress = 0
	main_simple_progress = ( 100 / len(inquiries) ) / int(deep_pars)

	# Запускаю selenium
	options = Options()
	# Запускаем драйвер без графической оболочки браузера
	options.headless = True
	# Убираем логирование в консоль
	options.add_argument('--log-level=3')
	# Инициализируем драйвер хром
	driver = webdriver.Chrome(chrome_options=options, executable_path='drivers/chromedriver.exe')


	for inquirie in inquiries:
		title_list = []
		description_list = []
		keywords_list = []
		h1_list = []
		h2_list = []
		h3_list = []
		for i in range(1, deep_pars + 1):

			# получаю страницу яндекс поиска
			q = 'https://yandex.ru/search/?text='+str(inquirie)+'&lr='+str(region)+'&p='+str(i)
			driver.get(q)
			soup = BeautifulSoup (driver.page_source, features="html.parser")
			links = []
			# обрабатываю полученную страницу
			for link in soup.select('.serp-item .organic__url'):
				# делаю сравнение со списком исключений для ссылок
				check_link = True
				for exception_val in exceptions:
					result = re.search(exception_val, link.attrs["href"])
					if result :
						check_link = False
						break
				# заполняю список собранными ссылками
				if check_link:
					links.append(link.attrs["href"])

			one_part_progress = round( main_simple_progress / len(links), 1 )

			for link in links:

				driver.get(link)
				soup_site = BeautifulSoup (driver.page_source, features="html.parser")

				if soup_site.title != None:
					title_list.append(soup_site.title.string)

				h3 = soup_site.find_all('h3')
				if h3 != None:
					for tag in h3:
						h3_list.append(tag.text)

				h2 = soup_site.find_all('h2')
				if h2 != None:
					for tag in h2:
						h2_list.append(tag.text)


				h1 = soup_site.find_all('h1')
				if h1 != None:
					for tag in h1:
						h1_list.append(tag.text)


				description = soup_site.find('meta', {'name':'description'})
				if description != None:
					description_list.append(description.get('content'))


				keywords = soup_site.find('meta', {'name':'keywords'})
				if keywords != None:
					keywords_list.append(keywords.get('content'))




				# создаю новую книгу
				workbook = openpyxl.Workbook()
				title_sheet = workbook.active

				# filename = datetime.datetime.today().strftime("%Y-%m-%d-%H-%M-%S")
				# выбираем активный лист и меняем ему название
				title_sheet.title = "title"

				if title_check.get() == True:
					i = 1
					for word in title_list:
						cellref = title_sheet.cell(row=i, column=1)
						cellref.value = word
						i = i + 1

				if h3_check.get() == True:
					# добавляю новую страницу
					h3_sheet = workbook.create_sheet('H3', 1)
					i = 1
					for word in h3_list: 
						cellref = h3_sheet.cell(row=i, column=1)
						cellref.value = word
						i = i + 1

				if h2_check.get() == True:
					# добавляю новую страницу
					h2_sheet = workbook.create_sheet('H2', 1)
					i = 1
					for word in h2_list:
						cellref = h2_sheet.cell(row=i, column=1)
						cellref.value = word
						i = i + 1

				# добавляю новую страницу
				if h1_check.get() == True:
					h1_sheet = workbook.create_sheet('H1', 1)
					i = 1
					for word in h1_list:
						cellref = h1_sheet.cell(row=i, column=1)
						cellref.value = word
						i = i + 1

				if keywords_check.get() == True:
					# добавляю новую страницу
					keywords_sheet = workbook.create_sheet('Keywords', 1)
					i = 1
					for word in keywords_list:
						cellref = keywords_sheet.cell(row=i, column=1)
						cellref.value = word
						i = i + 1

				if desc_check.get() == True:
					# добавляю новую страницу
					description_sheet = workbook.create_sheet('Description', 1)
					i = 1
					for word in description_list:
						cellref = description_sheet.cell(row=i, column=1)
						cellref.value = word
						i = i + 1

				# сохраняю данные в exel
				workbook.save(filename = inquirie+'.xlsx')

				progress = progress + one_part_progress
				print('Прогресс '+ str(progress))
				barVar.set(round(progress))
				root.update()
	

	button_1.config(state="normal")
	# button_2.config(state="disabled")
	barVar.set(100)
	root.update()

	print('Done')

# ---------------------------------------------------------------------------------
# рисую интерфейс
root = Tk()

root.title("Парсер мета данных сайтов по запросам")
root.geometry('400x450')
root.resizable(width=False, height=False)


frame_1 = Frame()
frame_2 = Frame()
frame_3 = Frame()
frame_4 = Frame()
frame_5 = Frame()
frame_6 = Frame()
frame_7 = Frame()
frame_8 = Frame()

frame_1.pack()
frame_2.pack()
frame_3.pack()
frame_4.pack()
frame_5.pack()
frame_6.pack()
frame_7.pack()
frame_8.pack()

lable_1 = Label(frame_1, text="Что собрать:")
lable_1.pack()

h1_check = BooleanVar()
h1_check.set(1)
c4 = Checkbutton(frame_1, text="h1", variable=h1_check, onvalue=1, offvalue=0)
c4.pack(side=LEFT)

title_check = BooleanVar()
title_check.set(1)
c1 = Checkbutton(frame_1, text="title", variable=title_check, onvalue=1, offvalue=0)
c1.pack(side=RIGHT)

h2_check = BooleanVar()
h2_check.set(1)
c5 = Checkbutton(frame_2, text="h2", variable=h2_check, onvalue=1, offvalue=0)
c5.pack(side=LEFT)

desc_check = BooleanVar()
desc_check.set(1)
c2 = Checkbutton(frame_2, text="desc", variable=desc_check, onvalue=1, offvalue=0)
c2.pack(side=RIGHT)

h3_check = BooleanVar()
h3_check.set(1)
c6 = Checkbutton(frame_3, text="h3", variable=h3_check, onvalue=1, offvalue=0)
c6.pack(side=LEFT)

keywords_check = BooleanVar()
keywords_check.set(1)
c3 = Checkbutton(frame_3, text="keys", variable=keywords_check, onvalue=1, offvalue=0)
c3.pack(side=RIGHT)

lable_2 = Label(frame_4, text="Регион:")
lable_2.pack()

combo = Combobox(frame_4)  
combo['values'] = (
	'Республика Татарстан', 
	'Казань', 
	'Москва и Московская область', 
	'Москва', 
	'Набережные Челны', 
	'Нижнекамск', 
	'Санкт-Петербург и Ленинградская область', 
	'Санкт-Петербург', 
	'Планета Земля', 
	'Без учета региона', 
	'Европа', 'СНГ', 
	'Универсальное', 
	'Азия', 'Россия', 
	'Северо-Западный федеральный округ', 
	'Калининградская область', 
	'Калининград', 
	'Мурманская область', 
	'Мурманск', 
	'Республика Карелия', 
	'Петрозаводск', 
	'Псковская область', 
	'Псков', 
	'Великие Луки', 
	'Новгородская область', 
	'Великий Новгород', 
	'Центральный федеральный округ', 
	'Тверская область', 
	'Тверь', 
	'Смоленская область', 
	'Смоленск', 
	'Брянская область', 
	'Брянск', 
	'Калужская область', 
	'Калуга', 
	'Обнинск', 
	'Курская область', 
	'Курск', 
	'Орловская область', 
	'Орёл', 
	'Тульская область', 
	'Тула', 
	'Долгопрудный', 
	'Дубна', 
	'Зеленоград', 
	'Пущино', 
	'Белгородская область', 
	'Белгород', 
	'Липецкая область', 
	'Липецк', 
	'Ярославская область', 
	'Ярославль', 
	'Владимирская область', 
	'Владимир', 
	'Александров', 
	'Гусь-Хрустальный', 
	'Муром', 
	'Ивановская область', 
	'Иваново', 
	'Рязанская область', 
	'Рязань', 
	'Тамбовская область', 
	'Тамбов', 
	'Воронежская область', 
	'Воронеж', 
	'Южный федеральный округ', 
	'Ростовская область', 
	'Ростов-на-Дону', 
	'Шахты', 
	'Таганрог', 
	'Новочеркасск', 
	'Волгодонск', 
	'Краснодарский край', 
	'Краснодар', 
	'Анапа', 
	'Новороссийск', 
	'Сочи', 
	'Туапсе', 
	'Геленджик', 
	'Армавир', 
	'Ейск', 
	'Республика Адыгея', 
	'Майкоп', 
	'Карачаево-Черкесская республика', 
	'Черкесск', 
	'Кабардино-Балкарская республика', 
	'Нальчик', 
	'Северная Осетия', 
	'Владикавказ', 
	'Республика Ингушетия', 
	'Чеченская республика', 
	'Грозный', 
	'Республика Дагестан', 
	'Махачкала', 
	'Ставропольский край', 
	'Ставрополь', 
	'Каменск-Шахтинский', 
	'Пятигорск', 
	'Минеральные Воды', 
	'Ессентуки', 
	'Кисловодск', 
	'Республика Калмыкия', 
	'Элиста', 
	'Астраханская область', 
	'Астрахань', 
	'Волгоградская область', 
	'Волгоград', 
	'Поволжье', 
	'Саратовская область', 
	'Саратов', 
	'Жигулевск', 
	'Балаково', 
	'Пензенская область', 
	'Пенза', 
	'Республика Мордовия', 
	'Саранск', 
	'Ульяновская область', 
	'Ульяновск', 
	'Самарская область', 
	'Самара', 
	'Тольятти', 
	'Сызрань', 
	'Чувашская республика', 
	'Чебоксары', 
	'Республика Марий Эл', 
	'Йошкар-Ола', 
	'Нижегородская область', 
	'Нижний Новгород', 
	'Саров', 
	'Кировская область', 
	'Киров', 
	'Костромская область', 
	'Кострома', 
	'Вологодская область', 
	'Вологда', 
	'Архангельская область', 
	'Архангельск', 
	'Северодвинск', 
	'Ненецкий автономный округ', 
	'Республика Коми', 
	'Сыктывкар', 
	'Удмуртская республика', 
	'Ижевск', 
	'Пермский край', 
	'Пермь', 
	'Республика Башкортостан', 
	'Уфа', 
	'Нефтекамск', 
	'Салават', 
	'Стерлитамак', 
	'Оренбургская область', 
	'Оренбург', 
	'Дзержинск', 
	'Урал', 
	'Челябинская область', 
	'Челябинск', 
	'Магнитогорск', 
	'Снежинск', 
	'Курганская область', 
	'Курган', 
	'Свердловская область', 
	'Екатеринбург', 
	'Каменск-Уральский', 
	'Нижний Тагил', 
	'Новоуральск', 
	'Первоуральск', 
	'Тюменская область', 
	'Тюмень', 
	'Тобольск', 
	'Ханты-Мансийский автономный округ', 
	'Ханты-Мансийск', 
	'Сургут', 
	'Нижневартовск', 
	'Сибирь', 
	'Омская область', 
	'Омск', 
	'Новосибирская область', 
	'Новосибирск', 
	'Бердск', 
	'Томская область', 
	'Томск', 
	'Ямало-Ненецкий автономный округ', 
	'Салехард', 
	'Алтайский край', 
	'Барнаул', 
	'Бийск', 
	'Рубцовск', 
	'Республика Алтай', 
	'Горно-Алтайск', 
	'Кемеровская область', 
	'Кемерово', 
	'Междуреченск', 
	'Новокузнецк', 
	'Прокопьевск', 
	'Республика Хакасия', 
	'Абакан', 
	'Республика Тыва', 
	'Кызыл', 
	'Красноярский край', 
	'Красноярск', 
	'Ачинск', 
	'Норильск', 
	'Железногорск', 
	'Иркутская область', 
	'Иркутск', 
	'Братск', 
	'Республика Бурятия', 
	'Улан-Удэ', 
	'Забайкальский край', 
	'Чита', 
	'Дальневосточный федеральный округ', 
	'Республика Саха (Якутия)', 
	'Якутск', 
	'Амурская область', 
	'Благовещенск', 
	'Еврейская автономная область', 
	'Биробиджан', 
	'Приморский край', 
	'Владивосток', 
	'Находка', 
	'Уссурийск', 
	'Чукотский автономный округ', 
	'Анадырь', 
	'Камчатский край', 
	'Петропавловск-Камчатский', 
	'Магаданская область', 
	'Магадан', 
	'Сахалинская область', 
	'Южно-Сахалинск', 
	'Хабаровский край', 
	'Хабаровск', 
	'Комсомольск-на-Амуре'
)
combo.current(1)  # установите вариант по умолчанию  
combo.pack()

lable_21 = Label(frame_5, text="Глубина парсинга:")
lable_21.pack()

spin = Spinbox(frame_5, from_=0, to=10, width=5, textvariable=1)
spin.pack()

lable_3 = Label(frame_6, text="Запросы:")
lable_3.pack()

text = Text(frame_6, width=38, height=12, bg="lightblue", fg='black', wrap=WORD)
text.pack()

button_1 = Button(frame_8, text="Собрать данные")
button_1.pack(side=RIGHT, pady=5)
button_1.bind("<Button-1>", output)

# button_2 = Button(frame_7, text="Остановить")
# button_2.pack(side=LEFT)
# button_2.bind("<Button-1>", stop_pars)
# button_2.config(state="disabled")

barVar = DoubleVar()
barVar.set(0)
progress_bar = Progressbar(frame_7, orient = HORIZONTAL, length = 300, variable=barVar, mode = 'determinate')
progress_bar.pack(pady=5)

root.event_add('<<Paste>>', '<Control-igrave>')
root.event_add("<<Copy>>", "<Control-ntilde>")

root.mainloop()